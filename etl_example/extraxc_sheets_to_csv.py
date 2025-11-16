import openpyxl
import pandas as pd
import numpy as np
import os


'''Extracts all sheets from two Excel workbooks (location_a and location_b),
cleans the data, and saves each sheet as a separate CSV.

- Loads XLSX files containing multiple sheets of raw data
- Detects and replaces merged cells in the price column with "NO_VALUE"
- Forward-fills empty date/group columns (merged cells are present in xlsx)
- Standardizes column names: PriceOrVoucher and SourceInfo
- Saves each cleaned sheet as CSV'''


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

data_sources = {
    "location_a": os.path.join(BASE_DIR, "data", "location_a", "raw", "source_file_a.xlsx"),
    "location_b": os.path.join(BASE_DIR, "data", "location_b", "raw", "source_file_b.xlsx"),
}

COLUMN_PRICE = "PriceOrVoucher"
COLUMN_INFO = "SourceInfo"

def process_location(label: str, xlsx_path: str):
    if not os.path.exists(xlsx_path):
        print(f"File not found for {label}: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames

    output_folder = os.path.join(BASE_DIR, "data", label, "processed")
    os.makedirs(output_folder, exist_ok=True)

    print(f"\nProcessing {label.upper()} ({len(sheet_names)} sheets)...")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        try:
            price_col_idx = headers.index(COLUMN_PRICE)
        except ValueError:
            print(f"'{COLUMN_PRICE}' not found in sheet: {sheet_name}")
            continue

        merged_to_replace = set()
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col - 1 == price_col_idx:
                for row in range(merged_range.min_row + 1, merged_range.max_row + 1):
                    merged_to_replace.add((row, merged_range.min_col))

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = []
            for j, cell in enumerate(row):
                if j == price_col_idx and (i, j + 1) in merged_to_replace:
                    values.append("NO_VALUE")
                else:
                    values.append(cell.value)
            rows.append(values)

        df = pd.DataFrame(rows, columns=headers)

        df.iloc[:, 0] = df.iloc[:, 0].replace(r'^\s*$', np.nan, regex=True).ffill()
        df.iloc[:, 1] = df.iloc[:, 1].replace(r'^\s*$', np.nan, regex=True).ffill()

        if COLUMN_INFO in df.columns:
            df[COLUMN_INFO] = df[COLUMN_INFO].replace(r'^\s*$', np.nan, regex=True).ffill()

        safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in sheet_name)
        csv_path = os.path.join(output_folder, f"{safe_name}.csv")

        df.to_csv(csv_path, index=False, encoding="utf-8")
        print(f"Saved: {csv_path}")

    print(f"Finished processing: {label.upper()}")

def main():
    for label, path in data_sources.items():
        process_location(label, path)
    print("\nAll processing complete.")

if __name__ == "__main__":
    main()
